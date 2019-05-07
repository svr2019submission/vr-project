#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import csv

import openpyxl

from collections import Counter


"""
    Given a xlsx worksheet, read the work_books and gather the informations about the smells
"""
def count_smels(work_sheet, non_vr_filter = None):
    
    #store the results
    smells_results = {}

    #target workbooks -> should extract information only os this workbooks
    target_workbooks = ['_ArchSMells', '_AbsSMells', '_EncSMells', '_ModSMells', '_HieSMells', '_ImpSmells', '_CodeClones']
    
    #headers names
    headers_list = ['Architecture smell', 'Design smell', 'Implementation smell', 'Clone-set Serial No']

    #a list with the name of the work_books
    work_books_list = work_sheet.sheetnames
    
    for work_book in work_books_list:
        
        #avoid unnecessary work books
        if not any(x in work_book for x in target_workbooks): #check sheets the contains any of the listed words
            continue

        #get a valid work book
        sheet = work_sheet[work_book]
        dict_filter = {}

        for col in sheet.iter_cols(min_row=1, max_col=1):
            
            #control which line is beeing read
            cell_id = 1

            for cell in col:

                row_content = sheet[cell_id]
                row_content = [str(x.value) for x in row_content if x.value]
                row_content = ' '.join(row_content) #convert the list into a single string to find any occurrance in a easier way

                #get the value of the cell 
                smell_name = cell.value

                #avoid headers
                if smell_name in headers_list:
                    continue               
                
                #handle code duplucation identifier
                if isinstance(smell_name, int) or smell_name is None:
                    smell_name = 'Clone'

                #dont need to filter                
                if not non_vr_filter:

                    #first occurrence
                    if not smell_name in smells_results.keys():
                        smells_results[smell_name] = 1
                    
                    #increment the number of occurrence
                    else:
                        smells_results[smell_name] += 1

                #need to filter before count
                else:
                    
                    #first occurrence
                    if not smell_name in smells_results.keys():
                        
                        #check if its a tested class
                        for class_name in non_vr_filter:
                            
                            class_name = str(class_name.strip()) #remove spaces and \n

                            #check row content
                            if class_name in row_content:

                                #count the smell
                                smells_results[smell_name] = 1
                                
                                #log the class and smell information
                                if not class_name in dict_filter.keys():
                                    dict_filter[class_name] = [1, [row_content]]
                                
                                #increment
                                else:
                                    temp        = dict_filter[class_name]
                                    counter     = temp[0] + 1
                                    new_content = temp[1] + [row_content]  
                                    dict_filter[class_name] = [counter, new_content]

                                #don't need to check the other classes
                                continue

                    #Wasn't first smell occurrence. Will increment the number of occurrences
                    else:

                        #check wich class the smell is related
                        for class_name in non_vr_filter:
                            class_name = str(class_name.strip()) #remove spaces and \n

                            #is this class?
                            if class_name in row_content:

                                #count the smell
                                smells_results[smell_name] = 1

                                #log the class and smell information
                                if not class_name in dict_filter.keys():             #first ocurrance
                                    dict_filter[class_name] = [1, [row_content]]
                                
                                else:
                                    temp        = dict_filter[class_name]
                                    counter     = temp[0] + 1
                                    new_content = temp[1] + [row_content]  
                                    dict_filter[class_name] = [counter, new_content]

                #update cell_id for the next row
                cell_id = cell_id + 1
    
    #order the keys by number of occurrance
    d_keys = sorted(dict_filter.keys()) 
    

    #print the dict
    #for key in d_keys:
        #print("{}: {}".format(key, dict_filter[key]))                    

    return smells_results, total


#python3 required
if __name__ == '__main__':

    #gather results
    results = dict()

    #base dir
    base_dir = os.getcwd()
    
    #Check python version
    if sys.version_info[0] < 3:
        raise Exception("Python 3 or a more recent version is required.")
        #run -> python3 smells_checker.py >> text.log    

	#choices
    print("1 - Run on VR projects")
    print("2 - Run on Non-VR projects")

    #read the option
    switch = int(input("Chose a target: "))
    
    #execute according to the option
    if switch not in [1, 2]:
        print("W: Invalid option!")


    repositories_vr_path  = base_dir + "/results/full_results/smells/vr"
    repositories_non_vr_path  = base_dir + "/results/full_results/smells/non-vr"

    #non-vr filter
    non_vr_filter_file = "./results/sampled_results/non_vr_tested_classes.csv"

    #convert the csv into a list
    non_vr_filter = []
    with open(non_vr_filter_file, 'rb') as f:
        for line in f:
            line = line.decode("utf-8").strip() #convert byte to string and remove \n from the string
            non_vr_filter.append(line)

    #check wich repositories will analize
    if switch == 1:
        repositories_path = repositories_vr_path
    else:
        repositories_path = repositories_non_vr_path

    #update the directory
    os.chdir(repositories_path)
    total = 0
    #check every file
    for root, dirs, files in os.walk(".", topdown=False):
        for filename in files:

            if '.xlsx' in filename:               
                
                result = dict()
                           
                #open a xlsx file
                work_sheet = openpyxl.load_workbook(filename)

                #decide wheter will filter the results or not
                if switch == 1:
                    result = count_smels(work_sheet) #VR , count smells normaly
                else:
                    result, temp = count_smels(work_sheet, non_vr_filter) #non-vr uses a filter, cont only smells of tested classes
                    total = total + temp
                #merge the results
                results = dict(Counter(results) + Counter(result))
    
    #count the smeels
    print(results)